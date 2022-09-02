#!/usr/bin/env python3
# coding=utf-8
# Author: libin
"""
pip install aliyun-python-sdk-ecs
参考官方文档：
    https://help.aliyun.com/document_detail/92989.html
    https://next.api.aliyun.com/api-tools/sdk/Ecs?version=2014-05-26&language=python
"""
import os
import logging
import json
import pprint
import threading
import time
import subprocess
import tqdm
from concurrent import futures
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkecs.request.v20140526.DescribeImagesRequest import DescribeImagesRequest
from aliyunsdkecs.request.v20140526.RunInstancesRequest import RunInstancesRequest
from aliyunsdkecs.request.v20140526.DescribeInstancesRequest import DescribeInstancesRequest
from aliyunsdkecs.request.v20140526.RebootInstancesRequest import RebootInstancesRequest
from aliyunsdkecs.request.v20140526.DescribeInstanceStatusRequest import DescribeInstanceStatusRequest

import random_pwd
import deliver_pubkey

AK_ID = ''
AK_SECRET = ''
REGION_ID = 'cn-shenzhen'
VSWITCH_ID='xxx'
SECURITY_GROUP_ID='xxx'

SERVERS_INFO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'servers_info.xlsx')
ANSIBLE_INVENTORY = '/etc/ansible/hosts'
ANSIBLE_PLAYBOOK = '/etc/ansible/auto_install.yml'
ANSIBLE_SOME_ARGS = f'-i {ANSIBLE_INVENTORY}'
SYSTEM_INITIALIZE_SCRIPT = '/usr/local/scripts/system_initialize.sh'

# log_file = os.path.join(os.path.dirname(__file__), 'create_ecs.log')
# logger = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
# stream_handler = logging.StreamHandler()
# stream_handler.setFormatter(formatter)
# logger.addHandler(stream_handler)
# logger.setLevel(logging.INFO)

# create a client object for connection
client = AcsClient(AK_ID, AK_SECRET, REGION_ID)

lock = threading.Lock()


def create_instances(hostname, instance_type,
                     data_disk_size=100,
                     amount=1,
                     private_ip_address=None,
                     password=None,
                     image_id='centos_7_9_x64_20G_alibase_20220629.vhd',
                     instance_charge_type='PrePaid',
                     period_unit='Month',
                     period=12,
                     system_disk_category='cloud_essd',
                     system_disk_size=40,
                     data_disk_category='cloud_efficiency',
                     internet_charge_type='PayByTraffic',
                     internet_max_bandwidth_out=5,
                     unique_suffix=True,
                     vswitch_id=VSWITCH_ID,
                     security_group_id=SECURITY_GROUP_ID):
    request = RunInstancesRequest()

    request.set_InstanceName(hostname)
    request.set_HostName(hostname)
    request.set_InstanceType(instance_type)
    if data_disk_size:
        request.set_DataDisks([
            {
                'Size': data_disk_size,
                'Category': data_disk_category
            }
        ])
    request.set_Amount(amount)
    if private_ip_address:
        request.set_PrivateIpAddress(private_ip_address)
    password = password if password else random_pwd.mkpasswd()
    request.set_Password(password)
    request.set_ImageId(image_id)
    request.set_InstanceChargeType(instance_charge_type)
    if instance_charge_type == 'PrePaid':
        request.set_PeriodUnit(period_unit)
        request.set_Period(period)
    request.set_SystemDiskCategory(system_disk_category)
    request.set_SystemDiskSize(system_disk_size)
    request.set_InternetChargeType(internet_charge_type)
    request.set_InternetMaxBandwidthOut(internet_max_bandwidth_out)
    request.set_UniqueSuffix(unique_suffix)
    request.set_VSwitchId(vswitch_id)
    request.set_SecurityGroupId(security_group_id)

    response = _send_request(request)
    instance_ids = response.get('InstanceIdSets').get('InstanceIdSet')
    print(f'成功创建ECS实例，实例ID为：{instance_ids}')
    return instance_ids, password


def output_instances_info(instance_ids: list, password: str | None = None) -> list:
    """输出ECS实例部分信息到文件"""
    if is_instances_running(instance_ids):
        request = DescribeInstancesRequest()
        request.set_InstanceIds(instance_ids)
        response = _send_request(request)
        instances_list = response.get('Instances').get('Instance')

        # 输出ECS信息到excel文件
        instances_info = []
        public_ip_addresses_list = []
        for instance_detail in instances_list:
            instance_id = instance_detail.get('InstanceId')
            instance_name = instance_detail.get('InstanceName')
            cpu = instance_detail.get('Cpu')
            memory = instance_detail.get('Memory') / 1024
            os_name = instance_detail.get('OSName')
            public_ip_address = instance_detail.get('PublicIpAddress').get('IpAddress')[0]
            private_ip_address = instance_detail.get('VpcAttributes').get('PrivateIpAddress').get('IpAddress')[0]
            instances_info.append([instance_id, instance_name, public_ip_address, private_ip_address, os_name, cpu, memory, password])
            public_ip_addresses_list.append(public_ip_address)
        lock.acquire()
        if not os.path.isfile(SERVERS_INFO_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(['instance_id', 'instance_name', 'public_ip_address', 'private_ip_address', 'os_name', 'cpu', 'memory', 'password'])
        else:
            wb = load_workbook(SERVERS_INFO_FILE)
            ws = wb.active
        for instance_info in instances_info:
            ws.append(instance_info)
        wb.save(SERVERS_INFO_FILE)
        lock.release()
        print(f'ecs实例 {instance_ids} 信息已输出到文件{SERVERS_INFO_FILE}')
        return public_ip_addresses_list


def is_instances_running(instance_ids: list) -> bool:
    amount = len(instance_ids)
    running_amount = 0
    max_try = 10
    flag = False
    while running_amount < amount and max_try > 0:
        max_try -= 1
        running_amount = 0
        response = describe_instance_statuses(instance_ids)
        instance_status_list = response.get('InstanceStatuses').get('InstanceStatus')
        for instance_status in instance_status_list:
            if instance_status.get('Status') == "Running":
                running_amount += 1
        if running_amount == amount:
            flag = True
            break
        time.sleep(10)
    if flag:
        print(f'ECS实例 {instance_ids} is running.')
    else:
        print(f'ECS实例 {instance_ids} is not running.')
    return flag


def add_to_ansible(public_ip_addresses_list: list, password: str):
    """将ecs实例添加到内网ansible管理"""
    for public_ip_address in public_ip_addresses_list:
        subprocess.run(f'ssh-keygen -R {public_ip_address}', shell=True, text=True, capture_output=True)
        errs = deliver_pubkey.deliver_pubkey(public_ip_address, password)
        if errs:
            raise RuntimeError(errs)
        lock.acquire()
        with open(ANSIBLE_INVENTORY, 'r+') as f:
            if public_ip_address not in f.read():
                print(public_ip_address, file=f)
        lock.release()
        print(f'已将 {public_ip_address} 添加到ansible管理')


def system_initialize(public_ip_addresses_list: list):
    """系统初始化"""
    for public_ip_address in public_ip_addresses_list:
        cmd = f"ansible {public_ip_address} {ANSIBLE_SOME_ARGS} -m script -a '{SYSTEM_INITIALIZE_SCRIPT}'"
        p = subprocess.run(cmd, shell=True, text=True, capture_output=True)
        if p.returncode:
            raise RuntimeError(p.stderr)
        print(f'{public_ip_address} 已完成系统初始化')


def install_components(public_ip_addresses_list: list, components: list | tuple):
    """安装中间件"""
    for public_ip_address in public_ip_addresses_list:
        cmd = f'ansible-playbook {ANSIBLE_SOME_ARGS} -e host={public_ip_address} -t {",".join(components)} {ANSIBLE_PLAYBOOK}'
        p = subprocess.run(cmd, shell=True, text=True, capture_output=True)
        if p.returncode:
            raise RuntimeError(p.stderr)
        print(f'{public_ip_address} 已安装中间件：{", ".join(components)}')


def reboot_instances(instance_ids: list):
    """重启ecs实例"""
    request = RebootInstancesRequest()
    request.set_InstanceIds(instance_ids)
    response = _send_request(request)
    return response


def reboot_os(public_ip_addresses_list: list):
    """重启操作系统"""
    for public_ip_address in public_ip_addresses_list:
        cmd = f"ansible {public_ip_address} {ANSIBLE_SOME_ARGS} -m shell -a 'shutdown -r now'"
        p = subprocess.run(cmd, shell=True, text=True, capture_output=True)


def _send_request(request):
    request.set_accept_format('json')
    response = json.loads(client.do_action_with_exception(request).decode('utf-8'))
    return response


def describe_images():
    """查询可使用的镜像资源"""
    request = DescribeImagesRequest()
    request.set_Architecture("x86_64")
    request.set_OSType("linux")
    request.set_PageSize(20)
    response = _send_request(request)
    return response


def describe_instances(instance_ids: list):
    """查询一台或多台ECS实例的详细信息"""
    request = DescribeInstancesRequest()
    request.set_InstanceIds(instance_ids)
    response = _send_request(request)
    return response


def describe_instance_statuses(instance_ids: list):
    """查询一台或多台ECS实例的状态信息"""
    request = DescribeInstanceStatusRequest()
    request.set_InstanceIds(instance_ids)
    response = _send_request(request)
    return response


def pipeline(hostname: str, instance_type: str, components: list | tuple | None = None, **kwargs):
    instance_ids, password = create_instances(hostname, instance_type, **kwargs)
    public_ip_addresses_list = output_instances_info(instance_ids, password)
    add_to_ansible(public_ip_addresses_list, password)
    system_initialize(public_ip_addresses_list)
    if components:
        install_components(public_ip_addresses_list, components)


params_list = [
    {'hostname': 'gateway', 'instance_type': 'ecs.s6-c1m1.small', 'amount': 2, 'components': ['nginx']},
    {'hostname': 'service', 'instance_type': 'ecs.s6-c1m1.small', 'amount': 2, 'components': ['redis']}
]

with futures.ThreadPoolExecutor() as executor:
    to_do = {executor.submit(pipeline, **params): params.get('hostname') for params in params_list}
    done_iter = tqdm.tqdm(futures.as_completed(to_do), total=len(to_do))
    for future in done_iter:
        e = future.exception()
        if e:
            print(f'Error for [{to_do.get(future)}]: {e}')


# instance_ids, password = create_instances(hostname='wofwf', instance_type='ecs.s6-c1m1.small', instance_charge_type='PostPaid', data_disk_size=0, amount=2)

